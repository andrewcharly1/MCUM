"""
Deterministic workbook extractor for MCUM supervised workers.

This module gives MCUM a local, auditable first step for spreadsheet tasks:
turn an .xlsx workbook into bounded JSON before a model worker reasons about it.
"""

from __future__ import annotations

import argparse
from datetime import date, datetime, timezone
import json
from pathlib import Path
import sys
from typing import Any


EXTRACTOR_VERSION = "mcum-spreadsheet-extractor-v1"


def _cell_to_json(value: Any, *, max_chars: int) -> Any:
    if value is None or isinstance(value, (bool, int, float)):
        return value
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    text = str(value)
    if len(text) > max_chars:
        return text[: max(0, max_chars - 24)].rstrip() + "...[cell clipped]"
    return text


def _non_empty(value: Any) -> bool:
    return value is not None and str(value).strip() != ""


def _row_to_json(row: tuple[Any, ...], *, max_chars: int) -> list[Any]:
    return [_cell_to_json(value, max_chars=max_chars) for value in row]


def _first_non_empty_row(
    worksheet: Any,
    *,
    max_scan_rows: int,
    max_cols: int,
) -> tuple[int | None, tuple[Any, ...] | None]:
    row_limit = min(int(worksheet.max_row or 0), max_scan_rows)
    col_limit = min(int(worksheet.max_column or 0), max_cols)
    if row_limit <= 0 or col_limit <= 0:
        return None, None
    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=row_limit, max_col=col_limit, values_only=True),
        start=1,
    ):
        if any(_non_empty(value) for value in row):
            return row_index, tuple(row)
    return None, None


def _sample_rows_after_header(
    worksheet: Any,
    *,
    header_row: int | None,
    max_rows: int,
    max_cols: int,
    max_chars: int,
) -> list[dict[str, Any]]:
    if not header_row:
        return []
    row_limit = int(worksheet.max_row or 0)
    col_limit = min(int(worksheet.max_column or 0), max_cols)
    if row_limit <= header_row or col_limit <= 0:
        return []
    samples: list[dict[str, Any]] = []
    for row_index, row in enumerate(
        worksheet.iter_rows(
            min_row=header_row + 1,
            max_row=row_limit,
            max_col=col_limit,
            values_only=True,
        ),
        start=header_row + 1,
    ):
        if not any(_non_empty(value) for value in row):
            continue
        samples.append(
            {
                "row_index": row_index,
                "values": _row_to_json(tuple(row), max_chars=max_chars),
            }
        )
        if len(samples) >= max_rows:
            break
    return samples


def _scan_sheet_cells(
    worksheet: Any,
    *,
    max_scan_rows: int,
    max_cols: int,
) -> dict[str, Any]:
    row_limit = min(int(worksheet.max_row or 0), max_scan_rows)
    col_limit = min(int(worksheet.max_column or 0), max_cols)
    non_empty_cells = 0
    formula_cells: list[dict[str, Any]] = []
    if row_limit <= 0 or col_limit <= 0:
        return {"non_empty_cells_scanned": 0, "formula_cells_in_scan": []}
    for row in worksheet.iter_rows(min_row=1, max_row=row_limit, max_col=col_limit):
        for cell in row:
            value = cell.value
            if _non_empty(value):
                non_empty_cells += 1
            if isinstance(value, str) and value.startswith("="):
                formula_cells.append({"cell": cell.coordinate, "formula": value[:240]})
                if len(formula_cells) >= 25:
                    break
        if len(formula_cells) >= 25:
            break
    return {
        "non_empty_cells_scanned": non_empty_cells,
        "formula_cells_in_scan": formula_cells,
    }


def _merged_ranges_summary(worksheet: Any) -> dict[str, Any]:
    merged = getattr(worksheet, "merged_cells", None)
    ranges = list(getattr(merged, "ranges", []) or [])
    return {
        "count": len(ranges),
        "sample": [str(item) for item in ranges[:20]],
    }


def extract_workbook(
    source_path: str | Path,
    *,
    max_sheets: int = 20,
    max_rows: int = 25,
    max_cols: int = 30,
    max_scan_rows: int = 200,
    max_cell_chars: int = 180,
) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - depends on local install state.
        return {
            "status": "failure",
            "extractor": EXTRACTOR_VERSION,
            "error": "openpyxl_not_available",
            "detail": str(exc),
        }

    path = Path(source_path).expanduser().resolve()
    if not path.exists():
        return {
            "status": "failure",
            "extractor": EXTRACTOR_VERSION,
            "error": "source_not_found",
            "source_path": str(path),
        }
    if path.suffix.lower() != ".xlsx":
        return {
            "status": "failure",
            "extractor": EXTRACTOR_VERSION,
            "error": "unsupported_file_type",
            "source_path": str(path),
            "suffix": path.suffix,
        }

    warnings: list[str] = []
    file_stat = path.stat()
    workbook_values = load_workbook(path, read_only=True, data_only=True)
    workbook_formulas = load_workbook(path, read_only=True, data_only=False)
    try:
        sheet_names = list(workbook_values.sheetnames)
        selected_sheet_names = sheet_names[: max(0, int(max_sheets))]
        if len(selected_sheet_names) < len(sheet_names):
            warnings.append(f"sheet list clipped from {len(sheet_names)} to {len(selected_sheet_names)}")

        sheets: list[dict[str, Any]] = []
        for sheet_name in selected_sheet_names:
            values_ws = workbook_values[sheet_name]
            formulas_ws = workbook_formulas[sheet_name]
            header_row, header_values = _first_non_empty_row(
                values_ws,
                max_scan_rows=max_scan_rows,
                max_cols=max_cols,
            )
            scan = _scan_sheet_cells(
                formulas_ws,
                max_scan_rows=max_scan_rows,
                max_cols=max_cols,
            )
            sheets.append(
                {
                    "name": sheet_name,
                    "sheet_state": getattr(values_ws, "sheet_state", "visible"),
                    "max_row": int(values_ws.max_row or 0),
                    "max_column": int(values_ws.max_column or 0),
                    "detected_header_row": header_row,
                    "headers": _row_to_json(tuple(header_values or ()), max_chars=max_cell_chars),
                    "sample_rows": _sample_rows_after_header(
                        values_ws,
                        header_row=header_row,
                        max_rows=max_rows,
                        max_cols=max_cols,
                        max_chars=max_cell_chars,
                    ),
                    "merged_ranges": _merged_ranges_summary(values_ws),
                    **scan,
                }
            )

        return {
            "status": "success",
            "extractor": EXTRACTOR_VERSION,
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "source_path": str(path),
            "file_name": path.name,
            "file_size_bytes": file_stat.st_size,
            "modified_at": datetime.fromtimestamp(file_stat.st_mtime, tz=timezone.utc).isoformat(),
            "limits": {
                "max_sheets": max_sheets,
                "max_rows_per_sheet": max_rows,
                "max_columns_per_sheet": max_cols,
                "max_scan_rows": max_scan_rows,
                "max_cell_chars": max_cell_chars,
            },
            "workbook": {
                "sheet_count": len(sheet_names),
                "sheet_names": sheet_names,
                "active_sheet": getattr(workbook_values.active, "title", None),
            },
            "sheets": sheets,
            "warnings": warnings,
        }
    finally:
        workbook_values.close()
        workbook_formulas.close()


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Extract a bounded JSON summary from an .xlsx workbook.")
    parser.add_argument("source_path", help="Path to the .xlsx workbook.")
    parser.add_argument("--output", help="Optional JSON output path.")
    parser.add_argument("--max-sheets", type=int, default=20)
    parser.add_argument("--max-rows", type=int, default=25)
    parser.add_argument("--max-cols", type=int, default=30)
    parser.add_argument("--max-scan-rows", type=int, default=200)
    parser.add_argument("--max-cell-chars", type=int, default=180)
    return parser


def main(argv: list[str] | None = None) -> int:
    args = _build_parser().parse_args(argv)
    result = extract_workbook(
        args.source_path,
        max_sheets=args.max_sheets,
        max_rows=args.max_rows,
        max_cols=args.max_cols,
        max_scan_rows=args.max_scan_rows,
        max_cell_chars=args.max_cell_chars,
    )
    output_text = json.dumps(result, ensure_ascii=False, indent=2)
    if args.output:
        output_path = Path(args.output).expanduser().resolve()
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(output_text + "\n", encoding="utf-8")
    print(output_text)
    return 0 if result.get("status") == "success" else 2


if __name__ == "__main__":
    raise SystemExit(main())
